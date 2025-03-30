import streamlit as st
import pandas as pd
import os
import datetime
from tempfile import NamedTemporaryFile
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from io import BytesIO
import openai
import anthropic
import re
import time
import requests
import concurrent.futures
from mistralai import Mistral

# === Page Config ===
st.set_page_config(page_title="LLM PDF Analyzer", page_icon="📊", layout="centered")

# === Header ===
st.image("https://cdn-icons-png.flaticon.com/512/270/270798.png", width=60)
st.title("Générateur de Comparaison LLMs")
st.markdown("Uploadez votre fichier Excel → Le système analyse les questions avec les LLMs et génère un fichier de sortie automatiquement lors de deux sessions d’interrogations distincte.")

# === Upload File ===
uploaded_file = st.file_uploader("Importer un fichier Excel", type=[".xlsx"])

if uploaded_file:
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name


    # Config Clés API (remplacer par variables d'environnement ou st.secrets)
    
    openai_api_key = st.secrets["openai_api_key"]
    mistral_api_key = st.secrets["mistral_api_key"]
    claude_api_key = st.secrets["claude_api_key"]

    if not all([openai_api_key, mistral_api_key, claude_api_key]):
        st.error("Erreur: Une ou plusieurs clés API sont manquantes.")
        st.stop()

    sheets = pd.read_excel(tmp_path, sheet_name=None)

    # Fonctions d'extraction (reprises du code)
    def extract_correct_letters(text):
        pattern = r'([A-F])[\.\:\)]?(.*?)(?=\n[A-F][\.\:\)]|\Z)'
        matches = re.findall(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        correct_letters = []
        for letter, block in matches:
            if re.search(r'\bVrai\b', block, flags=re.IGNORECASE):
                correct_letters.append(letter.upper())
        return ''.join(correct_letters)

    def extract_short_answer(response_text):
        return response_text.strip().split('\n')[0]

    def query_openai(prompt, max_retries=5):
        for attempt in range(max_retries):
            try:
                response = openai.ChatCompletion.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": prompt}],
                    api_key=openai_api_key
                )
                return response["choices"][0]["message"]["content"].strip()
            except openai.error.RateLimitError:
                time.sleep(2 ** attempt)
            except Exception as e:
                time.sleep(2)
        return ""

    def query_mistral(prompt, max_retries=5):
        url = "https://api.mistral.ai/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {mistral_api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "mistral-small-latest",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 200
        }
        for attempt in range(max_retries):
            try:
                response = requests.post(url, json=payload, headers=headers)
                if response.status_code == 429:
                    time.sleep(2 ** attempt)
                    continue
                response.raise_for_status()
                result = response.json()
                return result.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            except Exception as e:
                time.sleep(2)
        return ""

    def query_claude(prompt, max_retries=5):
        client_claude = anthropic.Anthropic(api_key=claude_api_key)
        for attempt in range(max_retries):
            try:
                response = client_claude.messages.create(
                    model="claude-3-5-haiku-20241022",
                    max_tokens=200,
                    messages=[{"role": "user", "content": prompt}]
                )
                return response.content[0].text.strip()
            except anthropic.RateLimitError:
                time.sleep(2 ** attempt)
            except Exception as e:
                time.sleep(2)
        return ""

    def query_all_models(prompt):
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_gpt = executor.submit(query_openai, prompt)
            future_mistral = executor.submit(query_mistral, prompt)
            future_claude = executor.submit(query_claude, prompt)
            return future_gpt.result(), future_mistral.result(), future_claude.result()

    def process_session(sheets, session_name):
        results = []
        for sheet_name, df in sheets.items():
            df = df.dropna(subset=["QUESTION", "TYPE DE QUESTION"])
            for index, row in df.iterrows():
                question = row.get('QUESTION', '').strip()
                question_type = row.get('TYPE DE QUESTION', '').strip().upper()
                correct_answer = str(row.get('RÉPONSE', 'Non fourni')).strip()

                prompt = ""
                extracted_gpt = extracted_mistral = extracted_claude = "N/A"

                if question_type == "QCM":
                    choices = [str(row.get(f'RÉPONSE {i}', '')).strip()
                               for i in range(1, 7) if pd.notna(row.get(f'RÉPONSE {i}', ''))]
                    if choices:
                        letters = ['A', 'B', 'C', 'D', 'E', 'F'][:len(choices)]
                        formatted_choices = '\n'.join([
                            f"{letters[i]}. {choice}" for i, choice in enumerate(choices)
                        ])
                        prompt = f"""
                        Question : {question}
                        {formatted_choices}

                        Indique la ou les réponses correctes en précisant pour chaque lettre s'il s'agit de Vrai ou Faux.
                        """
                    else:
                        prompt = f"Question : {question}\nPropose 4 à 6 réponses possibles et indique lesquelles sont correctes."
                elif question_type in ["VF", "VRAI/FAUX", "VRAI - FAUX"]:
                    choix_vf = [str(row.get(f'RÉPONSE {i}', '')).strip()
                                for i in range(1, 3) if pd.notna(row.get(f'RÉPONSE {i}', ''))]
                    if choix_vf:
                        letters = ['A', 'B'][:len(choix_vf)]
                        formatted_choices = '\n'.join([
                            f"{letters[i]}. {choice}" for i, choice in enumerate(choix_vf)
                        ])
                        prompt = f"""
                        Question : {question}
                        {formatted_choices}
                        Pour chaque affirmation, indique si elle est vraie ou fausse.
                        """
                    else:
                        prompt = f"Question : {question}\nRéponds uniquement par 'Vrai' ou 'Faux' avec justification."
                elif question_type == "QROC":
                    prompt = f"Question : {question}\nFournis une réponse courte, précise et directement liée à la question."
                else:
                    continue

                response_gpt, response_mistral, response_claude = query_all_models(prompt)

                if question_type == "QCM":
                    extracted_gpt = extract_correct_letters(response_gpt)
                    extracted_mistral = extract_correct_letters(response_mistral)
                    extracted_claude = extract_correct_letters(response_claude)
                elif question_type in ["VF", "VRAI/FAUX", "VRAI - FAUX"]:
                    extracted_gpt = extract_correct_letters(response_gpt)
                    extracted_mistral = extract_correct_letters(response_mistral)
                    extracted_claude = extract_correct_letters(response_claude)
                elif question_type == "QROC":
                    extracted_gpt = extract_short_answer(response_gpt)
                    extracted_mistral = extract_short_answer(response_mistral)
                    extracted_claude = extract_short_answer(response_claude)

                results.append([
                    sheet_name, question_type, question, prompt.strip(),
                    response_gpt, extracted_gpt,
                    response_mistral, extracted_mistral,
                    response_claude, extracted_claude,
                    correct_answer
                ])

        return results
    
    # Process both sessions
    results_session1 = process_session(sheets, "Session 1")
    results_session2 = process_session(sheets, "Session 2")

    df_full1 = pd.DataFrame(results_session1, columns=[
        "Feuille", "Type de question", "Question", "Prompt",
        "Réponse GPT", "Réponse extraite GPT",
        "Réponse Mistral", "Réponse extraite Mistral",
        "Réponse Claude", "Réponse extraite Claude",
        "Réponse correcte"
    ])
    df_full2 = pd.DataFrame(results_session2, columns=df_full1.columns)

    df1 = df_full1[[
        "Feuille", "Question",
        "Réponse GPT", "Réponse extraite GPT",
        "Réponse Mistral", "Réponse extraite Mistral",
        "Réponse Claude", "Réponse extraite Claude",
        "Réponse correcte"
    ]]
    df2 = df_full2[df1.columns]

    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Session 1")
    ws1.append(["Réponse1 = A = 1, Réponse2 = B = 2, etc."])
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)
    ws2 = wb.create_sheet("Session 2")
    ws2.append(["Réponse1 = A = 1, Réponse2 = B = 2, etc."])
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)

    output = BytesIO()
    wb.save(output)
    st.success("\u2705 Traitement terminé. Vous pouvez télécharger le fichier ci-dessous.")
    st.download_button("🗃️ Télécharger le fichier Excel généré", data=output.getvalue(), file_name="results_comparaison_llms_sessions.xlsx")
